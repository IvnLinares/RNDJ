#!/usr/bin/env python3
"""
Script para crear una plantilla con ejemplos de datos más realistas
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear workbook
wb = openpyxl.Workbook()

# ==================== HOJA 1: EJEMPLO CON DATOS ====================
ws_ejemplo = wb.active
ws_ejemplo.title = "Ejemplo - 50 Participantes"

# Headers
headers = ["Nombre Completo", "NIS", "Email", "Rama", "Notas"]
ws_ejemplo.append(headers)

# Formato headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws_ejemplo[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Datos de ejemplo realistas - Caminantes (15-18)
caminantes = [
    ["Juan Carlos Pérez López", "NIS001", "juan.perez@scout.com", "Caminantes", ""],
    ["María José González García", "NIS002", "maria.gonzalez@scout.com", "Caminantes", "Alérgica a maní"],
    ["Carlos Martínez Rodríguez", "NIS003", "carlos.martinez@scout.com", "Caminantes", ""],
    ["Ana López Torres", "NIS004", "ana.lopez@scout.com", "Caminantes", "Vegetariana"],
    ["Pedro Sánchez Silva", "NIS005", "pedro.sanchez@scout.com", "Caminantes", ""],
    ["Diana Ramírez Flores", "NIS006", "diana.ramirez@scout.com", "Caminantes", "Alérgica a productos lácteos"],
    ["Roberto Fernández Díaz", "NIS007", "roberto.fernandez@scout.com", "Caminantes", ""],
    ["Laura Mendoza García", "NIS008", "laura.mendoza@scout.com", "Caminantes", ""],
    ["Felipe Gutiérrez López", "NIS009", "felipe.gutierrez@scout.com", "Caminantes", "Vegetariano"],
    ["Natalia Vargas Ruiz", "NIS010", "natalia.vargas@scout.com", "Caminantes", ""],
    
    # Rovers (18-22)
    ["Alejandro Cruz Jiménez", "NIS011", "alejandro.cruz@scout.com", "Rovers", ""],
    ["Sofía Herrera Martínez", "NIS012", "sofia.herrera@scout.com", "Rovers", "Alérgica a gluten"],
    ["Miguel Ángel Delgado López", "NIS013", "miguel.delgado@scout.com", "Rovers", ""],
    ["Catalina Castillo García", "NIS014", "catalina.castillo@scout.com", "Rovers", ""],
    ["Javier Moreno Rodríguez", "NIS015", "javier.moreno@scout.com", "Rovers", "Vegetariano"],
    ["Isabella Romero Pérez", "NIS016", "isabella.romero@scout.com", "Rovers", ""],
    ["David Valenzuela López", "NIS017", "david.valenzuela@scout.com", "Rovers", ""],
    ["Valentina Sepúlveda Flores", "NIS018", "valentina.sepulveda@scout.com", "Rovers", "Alérgica a mariscos"],
    ["Andrés Araya Silva", "NIS019", "andres.araya@scout.com", "Rovers", ""],
    ["Gabriela Reyes García", "NIS020", "gabriela.reyes@scout.com", "Rovers", ""],
    
    # Dirigente Joven (22-25)
    ["Lucas Núñez Martínez", "NIS021", "lucas.nunez@scout.com", "Dirigente Joven", ""],
    ["Emma Rojas García", "NIS022", "emma.rojas@scout.com", "Dirigente Joven", "Vegetariana"],
    ["Mateo Jiménez López", "NIS023", "mateo.jimenez@scout.com", "Dirigente Joven", ""],
    ["Olivia Contreras Ruiz", "NIS024", "olivia.contreras@scout.com", "Dirigente Joven", "Alérgica a maní"],
    ["Santiago Bravo Flores", "NIS025", "santiago.bravo@scout.com", "Dirigente Joven", ""],
    ["Martina Carrera García", "NIS026", "martina.carrera@scout.com", "Dirigente Joven", ""],
    ["Camilo Fuentes López", "NIS027", "camilo.fuentes@scout.com", "Dirigente Joven", ""],
    ["Florencia Medina Pérez", "NIS028", "florencia.medina@scout.com", "Dirigente Joven", "Vegana"],
    ["Cristian Yáñez Silva", "NIS029", "cristian.yanez@scout.com", "Dirigente Joven", ""],
    ["Beatriz Parra García", "NIS030", "beatriz.parra@scout.com", "Dirigente Joven", ""],
    
    # Dirigente (25+)
    ["Francisco González López", "NIS031", "francisco.gonzalez@scout.com", "Dirigente", ""],
    ["Claudia Morales García", "NIS032", "claudia.morales@scout.com", "Dirigente", ""],
    ["Ricardo Díaz Martínez", "NIS033", "ricardo.diaz@scout.com", "Dirigente", "Vegetariano"],
    ["Patricia Riquelme Flores", "NIS034", "patricia.riquelme@scout.com", "Dirigente", ""],
    ["Fernando Herrera López", "NIS035", "fernando.herrera@scout.com", "Dirigente", ""],
    ["Marta Sandoval García", "NIS036", "marta.sandoval@scout.com", "Dirigente", "Alérgica a medicamentos"],
    ["Enrique Ramírez Silva", "NIS037", "enrique.ramirez@scout.com", "Dirigente", ""],
    ["Rosa María Flores Pérez", "NIS038", "rosa.flores@scout.com", "Dirigente", ""],
    ["Alberto Reyes López", "NIS039", "alberto.reyes@scout.com", "Dirigente", ""],
    ["Verónica Castillo García", "NIS040", "veronica.castillo@scout.com", "Dirigente", ""],
]

# Agregar datos
data_fill = PatternFill(start_color="E7F0F7", end_color="E7F0F7", fill_type="solid")
data_font = Font(size=11)
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row_data in caminantes:
    ws_ejemplo.append(row_data)

# Formato para datos
for row_idx in range(2, len(caminantes) + 2):
    for col_idx in range(1, len(headers) + 1):
        cell = ws_ejemplo.cell(row=row_idx, column=col_idx)
        cell.fill = data_fill
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

# Ajustar anchos
ws_ejemplo.column_dimensions['A'].width = 25
ws_ejemplo.column_dimensions['B'].width = 12
ws_ejemplo.column_dimensions['C'].width = 28
ws_ejemplo.column_dimensions['D'].width = 18
ws_ejemplo.column_dimensions['E'].width = 25

# ==================== HOJA 2: ESTADÍSTICAS ====================
ws_stats = wb.create_sheet("Estadísticas")

# Título
title_cell = ws_stats['A1']
title_cell.value = "📊 ESTADÍSTICAS DEL EJEMPLO"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
ws_stats.merge_cells('A1:B1')
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_stats.row_dimensions[1].height = 30

# Datos de estadísticas
stats = [
    ("", ""),
    ("TOTAL DE PARTICIPANTES:", 40),
    ("", ""),
    ("POR RAMA:", ""),
    ("Caminantes (15-18)", 10),
    ("Rovers (18-22)", 10),
    ("Dirigente Joven (22-25)", 10),
    ("Dirigente (25+)", 10),
    ("", ""),
    ("NOTAS ESPECIALES:", ""),
    ("Con notas de alergia", 8),
    ("Vegetarianos/Veganos", 5),
    ("Sin notas especiales", 27),
]

for row_idx, (label, value) in enumerate(stats, start=3):
    ws_stats[f'A{row_idx}'] = label
    if value:
        ws_stats[f'B{row_idx}'] = value
    
    if "TOTAL" in label or "POR RAMA" in label or "NOTAS ESPECIALES" in label:
        ws_stats[f'A{row_idx}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws_stats[f'A{row_idx}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        if value:
            ws_stats[f'B{row_idx}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws_stats[f'B{row_idx}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

ws_stats.column_dimensions['A'].width = 30
ws_stats.column_dimensions['B'].width = 15

# Guardar
wb.save('EJEMPLO_FORO_PARTICIPANTES.xlsx')
print("✅ Plantilla con ejemplos creada: EJEMPLO_FORO_PARTICIPANTES.xlsx")
print(f"   📊 40 participantes de ejemplo")
print(f"   📈 Hoja de estadísticas")
