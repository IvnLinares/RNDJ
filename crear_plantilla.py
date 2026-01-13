#!/usr/bin/env python3
"""
Script para crear la plantilla XLSX del Foro Nacional de Jóvenes
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear workbook
wb = openpyxl.Workbook()

# ==================== HOJA 1: PLANTILLA ====================
ws_plantilla = wb.active
ws_plantilla.title = "Participantes"

# Headers
headers = ["Nombre Completo", "NIS", "Email", "Rama", "Notas"]
ws_plantilla.append(headers)

# Formato headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws_plantilla[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Datos de ejemplo
ejemplos = [
    ["Juan Carlos Pérez López", "NIS001", "juan.perez@example.com", "Caminantes", "Sin alergias conocidas"],
    ["María González Rodríguez", "NIS002", "maria.gonzalez@example.com", "Rovers", "Alérgico a maní"],
    ["Carlos Martínez Silva", "NIS003", "carlos.martinez@example.com", "Dirigente Joven", ""],
    ["Ana López Torres", "NIS004", "ana.lopez@example.com", "Dirigente", "Vegetariano"],
]

data_fill = PatternFill(start_color="E7F0F7", end_color="E7F0F7", fill_type="solid")
data_font = Font(size=11)
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row_data in ejemplos:
    ws_plantilla.append(row_data)

# Formato para datos
for row_idx in range(2, len(ejemplos) + 2):
    for col_idx in range(1, len(headers) + 1):
        cell = ws_plantilla.cell(row=row_idx, column=col_idx)
        cell.fill = data_fill
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

# Ajustar anchos de columnas
ws_plantilla.column_dimensions['A'].width = 25  # Nombre
ws_plantilla.column_dimensions['B'].width = 12  # NIS
ws_plantilla.column_dimensions['C'].width = 28  # Email
ws_plantilla.column_dimensions['D'].width = 18  # Rama
ws_plantilla.column_dimensions['E'].width = 25  # Notas

# Filas en blanco para llenar
for row_idx in range(len(ejemplos) + 2, len(ejemplos) + 12):
    for col_idx in range(1, len(headers) + 1):
        cell = ws_plantilla.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = data_alignment

# ==================== HOJA 2: INSTRUCCIONES ====================
ws_info = wb.create_sheet("Instrucciones")

# Título
title_cell = ws_info['A1']
title_cell.value = "📋 INSTRUCCIONES DE IMPORTACIÓN - Foro Nacional de Jóvenes"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
ws_info.merge_cells('A1:C1')
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_info.row_dimensions[1].height = 30

# Contenido
instructions = [
    ("", ""),
    ("CAMPOS REQUERIDOS:", ""),
    ("1. Nombre Completo", "Nombre y apellido del participante (ej: Juan Pérez)"),
    ("2. NIS", "Número de Identificación Scout, único (ej: NIS001)"),
    ("3. Email", "Email válido del participante (ej: juan@example.com)"),
    ("4. Rama", "Una de: Caminantes, Rovers, Dirigente Joven, Dirigente"),
    ("5. Notas", "Campo opcional - Notas adicionales del participante"),
    ("", ""),
    ("RAMAS DISPONIBLES:", ""),
    ("🏔️  Caminantes", "Jóvenes de 15-18 años"),
    ("🧭 Rovers", "Jóvenes de 18-22 años"),
    ("⚜️  Dirigente Joven", "Jóvenes dirigentes de 22-25 años"),
    ("🎖️  Dirigente", "Dirigentes de 25+ años"),
    ("", ""),
    ("IMPORTANTE:", ""),
    ("✅ Los campos: Nombre, NIS y Email son OBLIGATORIOS", ""),
    ("✅ El NIS debe ser ÚNICO para cada participante", ""),
    ("✅ El Email debe ser ÚNICO para cada participante", ""),
    ("✅ Las Ramas EXACTO como aparece en la lista", ""),
    ("✅ Máximo 1000 registros por importación", ""),
    ("", ""),
    ("EJEMPLO DE USO:", ""),
    ("Nombre Completo", "María González López"),
    ("NIS", "NIS002"),
    ("Email", "maria@example.com"),
    ("Rama", "Rovers"),
    ("Notas", "Alérgica a maní"),
]

for row_idx, (col_a, col_b) in enumerate(instructions, start=3):
    ws_info[f'A{row_idx}'] = col_a
    ws_info[f'B{row_idx}'] = col_b
    
    if "CAMPOS REQUERIDOS" in col_a or "RAMAS DISPONIBLES" in col_a or "IMPORTANTE" in col_a or "EJEMPLO DE USO" in col_a:
        ws_info[f'A{row_idx}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws_info[f'A{row_idx}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_info.merge_cells(f'A{row_idx}:B{row_idx}')
    elif col_a.startswith(("1.", "2.", "3.", "4.", "5.", "✅")):
        ws_info[f'A{row_idx}'].font = Font(bold=True, size=10)
        ws_info[f'A{row_idx}'].alignment = Alignment(wrap_text=True, vertical="top")
        ws_info[f'B{row_idx}'].alignment = Alignment(wrap_text=True, vertical="top")
    elif col_a.startswith(("🏔️", "🧭", "⚜️", "🎖️")):
        ws_info[f'A{row_idx}'].font = Font(size=10, bold=True)
        ws_info[f'B{row_idx}'].font = Font(size=10)
        ws_info[f'A{row_idx}'].alignment = Alignment(wrap_text=True, vertical="top")
        ws_info[f'B{row_idx}'].alignment = Alignment(wrap_text=True, vertical="top")

ws_info.column_dimensions['A'].width = 35
ws_info.column_dimensions['B'].width = 40

# ==================== HOJA 3: RAMAS ====================
ws_ramas = wb.create_sheet("Ramas")

# Título
title_cell = ws_ramas['A1']
title_cell.value = "🏕️ RAMAS DEL SISTEMA"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
ws_ramas.merge_cells('A1:D1')
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_ramas.row_dimensions[1].height = 30

# Headers de tabla ramas
rama_headers = ["Rama", "Emoji", "Edad", "Descripción"]
ws_ramas.append(rama_headers)

for cell in ws_ramas[2]:
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Datos de ramas
ramas_data = [
    ["Caminantes", "🏔️", "15-18 años", "Jóvenes scouts en formación"],
    ["Rovers", "🧭", "18-22 años", "Jóvenes en transición"],
    ["Dirigente Joven", "⚜️", "22-25 años", "Dirigentes jóvenes en desarrollo"],
    ["Dirigente", "🎖️", "25+ años", "Dirigentes experimentados"],
]

for rama_data in ramas_data:
    ws_ramas.append(rama_data)

for row_idx in range(3, len(ramas_data) + 3):
    for col_idx in range(1, 5):
        cell = ws_ramas.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 4:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws_ramas.column_dimensions['A'].width = 18
ws_ramas.column_dimensions['B'].width = 10
ws_ramas.column_dimensions['C'].width = 12
ws_ramas.column_dimensions['D'].width = 35

# Guardar archivo
wb.save('PLANTILLA_FORO_PARTICIPANTES.xlsx')
print("✅ Plantilla creada: PLANTILLA_FORO_PARTICIPANTES.xlsx")
print(f"   📄 Hoja 1: Participantes (con ejemplos)")
print(f"   📋 Hoja 2: Instrucciones detalladas")
print(f"   🏕️  Hoja 3: Ramas disponibles")
